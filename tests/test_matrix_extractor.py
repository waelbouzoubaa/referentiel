from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from middleware.parser.grammar import MappingRule
from middleware.parser.matrix_extractor import _parse_tier_label, parse_matrix_file


# ─────────────────────────────────────────────────────────────────────────────
# Fixture — fichier Airisol synthétique
# ─────────────────────────────────────────────────────────────────────────────

def _make_airisol_xlsx(tmp_path: Path, nb_produits: int = 3) -> Path:
    """Crée un fichier Excel Airisol synthétique."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table 1"

    # Cartouche
    ws["E4"] = "2026-01-01"
    ws["J4"] = "2026-12-31"
    ws["A6"] = "Code article Ramery : 12345"

    # En-tête ligne 8 (tiers)
    ws.cell(row=8, column=7, value="0-500m²")
    ws.cell(row=8, column=9, value="500-1000m²")
    ws.cell(row=8, column=11, value=">1000m²")

    # En-tête ligne 9 (variantes)
    ws.cell(row=9, column=1, value="Famille")
    ws.cell(row=9, column=2, value="Sous-famille")
    ws.cell(row=9, column=3, value="Désignation")
    ws.cell(row=9, column=5, value="Épaisseur")
    ws.cell(row=9, column=6, value="R-value")
    ws.cell(row=9, column=7, value="ALU")
    ws.cell(row=9, column=8, value="BLANC")
    ws.cell(row=9, column=9, value="ALU")
    ws.cell(row=9, column=10, value="BLANC")
    ws.cell(row=9, column=11, value="ALU")
    ws.cell(row=9, column=12, value="BLANC")
    ws.cell(row=9, column=13, value="Franco")

    # Données à partir de la ligne 10
    for i in range(nb_produits):
        row = 10 + i
        ws.cell(row=row, column=1, value="Étanchéité")
        ws.cell(row=row, column=2, value="Sous-face")
        ws.cell(row=row, column=3, value=f"Produit Test {i+1}")
        ws.cell(row=row, column=5, value=float(40 + i * 10))
        ws.cell(row=row, column=6, value=float(1.0 + i * 0.5))
        # Prix ALU tier 1
        ws.cell(row=row, column=7, value=float(10 + i))
        # Prix BLANC tier 1
        ws.cell(row=row, column=8, value=float(9 + i))
        # Prix ALU tier 2
        ws.cell(row=row, column=9, value=float(8 + i))
        # Prix BLANC tier 2
        ws.cell(row=row, column=10, value=float(7 + i))
        # Prix ALU tier 3
        ws.cell(row=row, column=11, value=float(6 + i))
        # Prix BLANC tier 3
        ws.cell(row=row, column=12, value=float(5 + i))
        # Franco
        ws.cell(row=row, column=13, value=f"Franco à partir de {200 + i * 50}m²")

    path = tmp_path / "airisol_test.xlsx"
    wb.save(path)
    return path


def _make_airisol_rule() -> MappingRule:
    return MappingRule.model_validate({
        "supplier_code": "airisol",
        "mapping_version": 1,
        "sheet_match": "Table 1",
        "header_detection": {"mode": "explicit", "row": 9},
        "data_starts_row": 10,
        "extraction_mode": "matrix",
        "row_filter": {
            "must_have_value_in": ["C"],
            "must_have_value_in_any": ["G", "I", "K"],
        },
        "data_zone": {
            "rows": "10:31",
            "product_columns": "A:F",
            "price_matrix_columns": "G:L",
        },
        "product_columns": {
            "family": {"source_col": "A", "transform": "strip"},
            "subfamily": {"source_col": "B", "transform": "strip"},
            "designation": {"source_col": "C", "transform": "strip", "required": True},
            "supplier_product_code": {"source_col": "C", "transform": "strip", "required": True},
        },
        "attributes": [
            {"key": "epaisseur", "source_col": "E", "data_type": "decimal", "unit": "mm"},
            {"key": "r_value", "source_col": "F", "data_type": "decimal", "unit": "m².K/W"},
        ],
        "price_matrix": {
            "tier_axis": {"header_row": 8, "type": "quantity_range", "fallback_unit": "m²"},
            "variant_axis": {"header_row": 9, "dimension_name": "couleur"},
            "column_groups": [
                {"columns": ["G", "H"], "tier_label": "0-500m²", "variants": ["ALU", "BLANC"]},
                {"columns": ["I", "J"], "tier_label": "500-1000m²", "variants": ["ALU", "BLANC"]},
                {"columns": ["K", "L"], "tier_label": ">1000m²", "variants": ["ALU", "BLANC"]},
            ],
            "price_type": "list",
            "currency": "EUR",
            "transform": "parse_decimal_fr",
        },
        "commercial_rules": [
            {
                "source_col": "M",
                "rule_type": "franco",
                "threshold_unit": "m²",
                "parse_pattern": "(\\d+)\\s*m²",
            }
        ],
        "file_metadata": {
            "validity_start": {"cell": "E4", "transform": "parse_date_iso"},
            "validity_end": {"cell": "J4", "transform": "parse_date_iso"},
            "client_article_code": {
                "regex": "Code article Ramery\\s*:\\s*(\\d+)",
                "in_cell": "A6",
            },
        },
        "gery_export": {
            "enabled": True,
            "flatten_strategy": "cartesian",
            "defaults": {"item_purchase_type": "Catalogue", "minimum_quantity": 1},
            "price_export_mapping": {"direct_unit_cost": "list"},
        },
    })


# ─────────────────────────────────────────────────────────────────────────────
# Tests — _parse_tier_label
# ─────────────────────────────────────────────────────────────────────────────

def test_parse_tier_label_range() -> None:
    mn, mx, unit = _parse_tier_label("0-500m²")
    assert mn == Decimal("0")
    assert mx == Decimal("500")
    assert unit == "m²"


def test_parse_tier_label_range_mid() -> None:
    mn, mx, unit = _parse_tier_label("500-1000m²")
    assert mn == Decimal("500")
    assert mx == Decimal("1000")


def test_parse_tier_label_gt() -> None:
    mn, mx, unit = _parse_tier_label(">1000m²")
    assert mn == Decimal("1000")
    assert mx is None
    assert unit == "m²"


# ─────────────────────────────────────────────────────────────────────────────
# Tests — parsing E2E
# ─────────────────────────────────────────────────────────────────────────────

def test_parse_airisol_nb_produits(tmp_path: Path) -> None:
    path = _make_airisol_xlsx(tmp_path, nb_produits=3)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)
    assert result.error_count == 0
    assert len(result.products) == 3


def test_parse_airisol_champs_produit(tmp_path: Path) -> None:
    path = _make_airisol_xlsx(tmp_path, nb_produits=2)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)

    p = result.products[0]
    assert p.supplier_code == "airisol"
    assert p.designation == "Produit Test 1"
    assert p.family == "Étanchéité"
    assert p.subfamily == "Sous-face"
    assert p.product_kind == "physical"


def test_parse_airisol_variantes(tmp_path: Path) -> None:
    path = _make_airisol_xlsx(tmp_path, nb_produits=1)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)

    p = result.products[0]
    assert len(p.variants) == 2
    codes = {v.variant_code for v in p.variants}
    assert codes == {"ALU", "BLANC"}
    dimensions = {v.variant_dimension for v in p.variants}
    assert dimensions == {"couleur"}


def test_parse_airisol_prix_par_variante(tmp_path: Path) -> None:
    """Chaque variante doit avoir 3 prix (un par palier)."""
    path = _make_airisol_xlsx(tmp_path, nb_produits=1)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)

    p = result.products[0]
    alu = next(v for v in p.variants if v.variant_code == "ALU")
    assert len(alu.prices) == 3

    tiers = sorted([(pr.tier_min_quantity, pr.tier_max_quantity) for pr in alu.prices])
    assert tiers[0] == (Decimal("0"), Decimal("500"))
    assert tiers[1] == (Decimal("500"), Decimal("1000"))
    assert tiers[2] == (Decimal("1000"), None)


def test_parse_airisol_valeur_prix(tmp_path: Path) -> None:
    path = _make_airisol_xlsx(tmp_path, nb_produits=1)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)

    p = result.products[0]
    alu = next(v for v in p.variants if v.variant_code == "ALU")
    # Tier 1 prix ALU = 10.0
    tier1 = next(pr for pr in alu.prices if pr.tier_min_quantity == Decimal("0"))
    assert tier1.amount == Decimal("10")
    assert tier1.price_type == "list"
    assert tier1.currency == "EUR"


def test_parse_airisol_attributs(tmp_path: Path) -> None:
    path = _make_airisol_xlsx(tmp_path, nb_produits=1)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)

    p = result.products[0]
    keys = {a.key for a in p.attributes}
    assert "epaisseur" in keys
    assert "r_value" in keys
    epaisseur = next(a for a in p.attributes if a.key == "epaisseur")
    assert epaisseur.unit == "mm"


def test_parse_airisol_commercial_rule(tmp_path: Path) -> None:
    path = _make_airisol_xlsx(tmp_path, nb_produits=1)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)

    p = result.products[0]
    assert len(p.commercial_rules) == 1
    cr = p.commercial_rules[0]
    assert cr.rule_type == "franco"
    assert cr.threshold_value == Decimal("200")
    assert cr.threshold_unit == "m²"


def test_parse_airisol_file_metadata(tmp_path: Path) -> None:
    path = _make_airisol_xlsx(tmp_path, nb_produits=1)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)

    meta = result.file_metadata
    assert meta.validity_start is not None
    assert meta.validity_end is not None


def test_parse_airisol_all_prices(tmp_path: Path) -> None:
    """all_prices() retourne 6 prix (3 tiers × 2 variantes) par produit."""
    path = _make_airisol_xlsx(tmp_path, nb_produits=1)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)

    p = result.products[0]
    assert len(p.all_prices()) == 6


def test_parse_airisol_ligne_sans_prix_ignoree(tmp_path: Path) -> None:
    """Une ligne avec colonnes G/I/K vides est exclue par row_filter."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table 1"
    # Ligne valide
    ws.cell(row=10, column=3, value="Produit valide")
    ws.cell(row=10, column=7, value=10.0)
    # Ligne sans prix → doit être ignorée
    ws.cell(row=11, column=3, value="Produit sans prix")
    path = tmp_path / "airisol_vide.xlsx"
    wb.save(path)
    rule = _make_airisol_rule()
    result = parse_matrix_file(path, rule)
    assert len(result.products) == 1
