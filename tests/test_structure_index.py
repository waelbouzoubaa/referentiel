"""Tests — reconnaissance de structure (empreinte d'en-tête exacte, sans IA)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from middleware.parser.grammar import HeaderDetection, MappingRule
from middleware.parser.pivot import FileMetadataPivot
from middleware.structure_index import (
    build_auto_validated_rule,
    extract_header_signature,
    find_matching_supplier,
    get_expected_file_metadata_fields,
    update_fingerprint,
    verify_file_metadata_presence,
)

_HEADERS = {"B": "Code article", "C": "Désignation", "D": "Prix"}


def _rule(row: int = 9) -> MappingRule:
    return MappingRule.model_validate({
        "supplier_code": "atlantic_scga_chauffage",
        "mapping_version": 1,
        "sheet_match": "auto",
        "header_detection": {"mode": "explicit", "row": row},
        "data_starts_row": row + 1,
        "extraction_mode": "table",
        "columns": {
            "supplier_product_code": {"source_col": "B", "required": True},
            "designation": {"source_col": "C", "required": True},
        },
        "gery_export": {"enabled": True, "flatten_strategy": "cartesian"},
    })


def _make_file(tmp_path: Path, name: str, header_row: int, headers: dict[str, str]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, text in headers.items():
        ws[f"{col}{header_row}"] = text
    ws[f"B{header_row + 1}"] = "CODE001"
    ws[f"C{header_row + 1}"] = "Un produit"
    path = tmp_path / name
    wb.save(path)
    return path


def _write_yaml(path: Path, rule: MappingRule) -> None:
    import io

    from ruamel.yaml import YAML
    yaml = YAML(typ="safe")
    stream = io.StringIO()
    yaml.dump(rule.model_dump(mode="json", exclude_none=True), stream)
    path.write_text(stream.getvalue(), encoding="utf-8")


@pytest.fixture(autouse=True)
def _isolated_index(tmp_path, monkeypatch):
    """Redirige le registre et le dossier suppliers vers des chemins temporaires."""
    import middleware.structure_index as si
    monkeypatch.setattr(si, "INDEX_FILE", tmp_path / "structure_index.json")
    (tmp_path / "suppliers").mkdir(exist_ok=True)
    monkeypatch.setattr(si, "CONFIG_DIR", tmp_path / "suppliers")
    yield


def test_meme_structure_reconnue_malgre_nom_different(tmp_path: Path) -> None:
    rule = _rule(row=9)
    original = _make_file(tmp_path, "original.xlsx", 9, _HEADERS)
    update_fingerprint("atlantic_scga_chauffage", original, rule)

    duplicate = _make_file(tmp_path, "AUTRE_FOURNISSEUR_v2.xlsx", 9, _HEADERS)
    assert find_matching_supplier(duplicate) == ("atlantic_scga_chauffage", 9)


def test_espace_et_casse_ignores(tmp_path: Path) -> None:
    rule = _rule(row=9)
    original = _make_file(tmp_path, "original.xlsx", 9, _HEADERS)
    update_fingerprint("atlantic_scga_chauffage", original, rule)

    variant_headers = {"B": "  code   article", "C": "désignation", "D": "PRIX"}
    variant = _make_file(tmp_path, "variant.xlsx", 9, variant_headers)
    assert find_matching_supplier(variant) == ("atlantic_scga_chauffage", 9)


def test_decalage_de_colonne_ne_matche_pas(tmp_path: Path) -> None:
    rule = _rule(row=9)
    original = _make_file(tmp_path, "original.xlsx", 9, _HEADERS)
    update_fingerprint("atlantic_scga_chauffage", original, rule)

    # Même libellés, mais décalés d'une colonne vers la droite (C/D/E au lieu de B/C/D)
    shifted_headers = {"C": "Code article", "D": "Désignation", "E": "Prix"}
    shifted = _make_file(tmp_path, "shifted.xlsx", 9, shifted_headers)
    assert find_matching_supplier(shifted) is None


def test_colonne_en_plus_ne_matche_pas(tmp_path: Path) -> None:
    rule = _rule(row=9)
    original = _make_file(tmp_path, "original.xlsx", 9, _HEADERS)
    update_fingerprint("atlantic_scga_chauffage", original, rule)

    extra_headers = {**_HEADERS, "E": "Stock"}
    extra = _make_file(tmp_path, "extra.xlsx", 9, extra_headers)
    assert find_matching_supplier(extra) is None


def test_aucune_empreinte_connue(tmp_path: Path) -> None:
    fichier = _make_file(tmp_path, "seul.xlsx", 9, {"B": "Code article", "C": "Désignation"})
    assert find_matching_supplier(fichier) is None


def test_extract_header_signature_mode_auto_retourne_none(tmp_path: Path) -> None:
    rule = _rule(row=9).model_copy(update={"header_detection": HeaderDetection(mode="auto")})
    fichier = _make_file(tmp_path, "f.xlsx", 9, {"B": "Code article", "C": "Désignation"})
    assert extract_header_signature(fichier, rule) is None


# ─────────────────────────────────────────────────────────────────────────────
# Présence des champs file_metadata (cartouche) — jamais les valeurs
# ─────────────────────────────────────────────────────────────────────────────

def test_verify_file_metadata_presence(tmp_path: Path) -> None:
    from datetime import date

    rule = _rule(row=9)
    original = _make_file(tmp_path, "original.xlsx", 9, _HEADERS)
    fm = FileMetadataPivot(validity_start=date(2026, 1, 1), ramery_generic_code="1750")
    update_fingerprint("atlantic_scga_chauffage", original, rule, fm)

    assert get_expected_file_metadata_fields("atlantic_scga_chauffage") == [
        "ramery_generic_code", "validity_start",
    ]

    # même champs présents, valeurs différentes (SIREN/dates différents = normal)
    ok = FileMetadataPivot(validity_start=date(2026, 6, 1), ramery_generic_code="9999")
    assert verify_file_metadata_presence("atlantic_scga_chauffage", ok) is True

    # un champ attendu (ramery_generic_code) absent cette fois → refusé
    incomplet = FileMetadataPivot(validity_start=date(2026, 6, 1))
    assert verify_file_metadata_presence("atlantic_scga_chauffage", incomplet) is False


def test_verify_file_metadata_presence_chaine_vide_compte_comme_absent(tmp_path: Path) -> None:
    """Certains lecteurs Excel renvoient '' (pas None) pour une cellule vraiment
    vide — régression constatée en test réel (2026-09-04) : ne doit pas compter
    comme "présent"."""
    rule = _rule(row=9)
    original = _make_file(tmp_path, "original.xlsx", 9, _HEADERS)
    fm = FileMetadataPivot(ramery_generic_code="1750")
    update_fingerprint("atlantic_scga_chauffage", original, rule, fm)

    vide = FileMetadataPivot(ramery_generic_code="")
    assert verify_file_metadata_presence("atlantic_scga_chauffage", vide) is False

    blanc = FileMetadataPivot(ramery_generic_code="   ")
    assert verify_file_metadata_presence("atlantic_scga_chauffage", blanc) is False


def test_verify_file_metadata_presence_aucun_champ_attendu(tmp_path: Path) -> None:
    """Sans file_metadata fourni à update_fingerprint, rien n'est exigé (rétro-compat)."""
    rule = _rule(row=9)
    original = _make_file(tmp_path, "original.xlsx", 9, _HEADERS)
    update_fingerprint("atlantic_scga_chauffage", original, rule)  # pas de file_metadata

    assert get_expected_file_metadata_fields("atlantic_scga_chauffage") == []
    assert verify_file_metadata_presence("atlantic_scga_chauffage", FileMetadataPivot()) is True


# ─────────────────────────────────────────────────────────────────────────────
# Recalcul de plage pour l'auto-validation (matrix / multi_table)
# ─────────────────────────────────────────────────────────────────────────────

def _matrix_rule(rows: str = "10:15") -> MappingRule:
    return MappingRule.model_validate({
        "supplier_code": "airisol_test",
        "mapping_version": 1,
        "sheet_match": "auto",
        "header_detection": {"mode": "explicit", "row": 9},
        "data_starts_row": 10,
        "extraction_mode": "matrix",
        "data_zone": {"rows": rows, "product_columns": "A:C", "price_matrix_columns": "D:D"},
        "product_columns": {
            "designation": {"source_col": "C", "required": True},
        },
        "price_matrix": {
            "tier_axis": {"header_row": 8, "detect_per_block": False},
            "variant_axis": {"header_row": 9, "dimension_name": "couleur"},
            "column_groups": [{"columns": ["D"], "tier_label": "0-500", "variants": ["ALU"]}],
        },
        "gery_export": {"enabled": True, "flatten_strategy": "cartesian"},
    })


def test_build_auto_validated_rule_matrix_etend_la_plage(tmp_path: Path) -> None:
    rule = _matrix_rule(rows="10:15")
    _write_yaml(tmp_path / "suppliers" / "airisol_test_v1.yaml", rule)

    # Nouveau fichier avec plus de lignes de produits que l'original (10:15)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["C9"] = "Désignation"
    for r in range(10, 21):
        ws[f"C{r}"] = f"Produit {r}"
        ws[f"D{r}"] = 100
    new_file = tmp_path / "nouveau.xlsx"
    wb.save(new_file)

    from middleware.parser.excel_reader import find_sheet, read_workbook
    sheets = read_workbook(new_file)
    _, sheet = find_sheet(sheets, "auto")
    expected_len = len(sheet)

    adapted = build_auto_validated_rule("airisol_test", new_file, new_header_row=9)
    assert adapted is not None
    start, end = adapted.data_zone.rows.split(":")
    assert start == "10"
    assert int(end) == expected_len  # étendu à la vraie taille, pas figé à 15


def test_build_auto_validated_rule_mode_auto_retourne_none(tmp_path: Path) -> None:
    rule = _matrix_rule().model_copy(update={"header_detection": HeaderDetection(mode="auto")})
    _write_yaml(tmp_path / "suppliers" / "airisol_test_v1.yaml", rule)
    new_file = _make_file(tmp_path, "nouveau.xlsx", 9, _HEADERS)
    assert build_auto_validated_rule("airisol_test", new_file, new_header_row=9) is None


def test_build_auto_validated_rule_yaml_introuvable(tmp_path: Path) -> None:
    new_file = _make_file(tmp_path, "nouveau.xlsx", 9, _HEADERS)
    assert build_auto_validated_rule("code_jamais_approuve", new_file, new_header_row=9) is None


def test_build_auto_validated_rule_decale_si_entete_ligne_differente(tmp_path: Path) -> None:
    """Reproduit un bug constaté en test réel (2026-09-04) : même en-tête, mais à
    une ligne différente dans le nouveau fichier (cartouche plus long au-dessus).
    Sans décalage, data_starts_row resterait figé sur l'ancienne position et la
    ligne d'en-tête elle-même serait lue comme une ligne de produit."""
    rule = _rule(row=1).model_copy(update={"data_starts_row": 2})
    _write_yaml(tmp_path / "suppliers" / "atlantic_scga_chauffage_v1.yaml", rule)

    # Dans le nouveau fichier, l'en-tête est à la ligne 3 (pas 1).
    new_file = _make_file(tmp_path, "nouveau.xlsx", 3, _HEADERS)
    adapted = build_auto_validated_rule("atlantic_scga_chauffage", new_file, new_header_row=3)
    assert adapted is not None
    assert adapted.header_detection.row == 3
    assert adapted.data_starts_row == 4  # décalé de +2, comme l'en-tête


def test_scenario_entete_decalee_pas_de_faux_produit(tmp_path: Path) -> None:
    """Bout en bout : structure identique mais en-tête à une ligne différente —
    le mapping adapté ne doit PAS lire la ligne d'en-tête comme un produit
    (régression du bug constaté en test réel 2026-09-04)."""
    rule = _rule(row=1).model_copy(update={"data_starts_row": 2})
    original = _make_file(tmp_path, "original.xlsx", 1, _HEADERS)
    update_fingerprint("atlantic_scga_chauffage", original, rule)
    _write_yaml(tmp_path / "suppliers" / "atlantic_scga_chauffage_v1.yaml", rule)

    nouveau = _make_file(tmp_path, "nouveau.xlsx", 3, _HEADERS)  # cartouche 2 lignes plus long
    match = find_matching_supplier(nouveau)
    assert match == ("atlantic_scga_chauffage", 3)

    adapted = build_auto_validated_rule(match[0], nouveau, match[1])
    assert adapted is not None
    assert adapted.data_starts_row == 4

    from middleware.parser.table_extractor import parse_table_file
    result = parse_table_file(nouveau, adapted)
    assert len(result.products) == 1
    assert result.products[0].supplier_product_code == "CODE001"


def _multi_table_rule() -> MappingRule:
    return MappingRule.model_validate({
        "supplier_code": "agenor_test",
        "mapping_version": 1,
        "sheet_match": "auto",
        "header_detection": {"mode": "explicit", "row": 1},
        "data_starts_row": 2,
        "extraction_mode": "multi_table",
        "product_kind": "service",
        "tables": [
            {"name": "table1", "zone": {"header_row": 1, "data_rows": "2:5", "cols": "A:B"},
             "layout": "bareme_1D"},
            {"name": "table2", "zone": {"header_row": 10, "data_rows": "11:15", "cols": "A:B"},
             "layout": "bareme_1D"},
        ],
        "gery_export": {"enabled": True, "flatten_strategy": "cartesian"},
    })


def test_build_auto_validated_rule_multi_table_borne_par_tableau_suivant(tmp_path: Path) -> None:
    rule = _multi_table_rule()
    _write_yaml(tmp_path / "suppliers" / "agenor_test_v1.yaml", rule)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Entête 1"
    for r in range(2, 9):  # plus de lignes que l'original (2:5) pour le 1er tableau
        ws[f"A{r}"] = f"ligne {r}"
    ws["A10"] = "Entête 2"
    for r in range(11, 26):  # plus de lignes que l'original (11:15) pour le 2e tableau
        ws[f"A{r}"] = f"ligne {r}"
    new_file = tmp_path / "nouveau.xlsx"
    wb.save(new_file)

    from middleware.parser.excel_reader import find_sheet, read_workbook
    sheets = read_workbook(new_file)
    _, sheet = find_sheet(sheets, "auto")
    expected_len = len(sheet)

    adapted = build_auto_validated_rule("agenor_test", new_file, new_header_row=1)
    assert adapted is not None
    t1_start, t1_end = adapted.tables[0].zone.data_rows.split(":")
    t2_start, t2_end = adapted.tables[1].zone.data_rows.split(":")
    assert t1_start == "2"
    assert int(t1_end) == 9  # header_row du tableau suivant (10) - 1, jamais deviné
    assert t2_start == "11"
    assert int(t2_end) == expected_len  # dernier tableau -> étendu à la fin de la feuille
