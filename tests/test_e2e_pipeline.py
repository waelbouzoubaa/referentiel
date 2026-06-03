"""Tests E2E — pipeline complet : parsing → delta → export Gery.

Ces tests couvrent le chemin doré sans base de données :
- fichier synthétique en mémoire
- parsing (table/matrix/multi_table)
- delta engine
- génération des fichiers Gery
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from middleware.delta.engine import ChangeType, compute_delta
from middleware.exporter.gery import generate_gery_exports
from middleware.parser.grammar import MappingRule
from middleware.parser.pivot import ProductPivot
from middleware.parser.table_extractor import compute_business_hash, parse_table_file
from middleware.parser.matrix_extractor import parse_matrix_file


# ─────────────────────────────────────────────────────────────────────────────
# Helpers — fichiers synthétiques
# ─────────────────────────────────────────────────────────────────────────────

def _atlantic_rule() -> MappingRule:
    return MappingRule.model_validate({
        "supplier_code": "atlantic_scga_chauffage",
        "mapping_version": 1,
        "sheet_match": "Atlantic 2026",
        "header_detection": {"mode": "explicit", "row": 9},
        "data_starts_row": 10,
        "extraction_mode": "table",
        "row_filter": {
            "must_have_value_in": ["B"],
            "exclude_if_starts_with": ["TARIF GROUPE"],
        },
        "columns": {
            "supplier_product_code": {"source_col": "B", "transform": ["strip", "to_uppercase"], "required": True},
            "designation": {"source_col": "C", "transform": "strip", "required": True},
            "family": {"constant": "Chauffage électrique"},
        },
        "prices": [
            {"type": "public", "source_col": "E", "transform": "parse_decimal_fr"},
            {"type": "installer", "source_col": "F", "transform": "parse_decimal_fr"},
        ],
        "attributes": [
            {"key": "quantity_pack", "source_col": "D", "data_type": "integer"},
        ],
        "file_metadata": {
            "validity_start": {"cell": "C4", "transform": "parse_date_iso"},
            "validity_end": {"cell": "C5", "transform": "parse_date_iso"},
        },
        "gery_export": {
            "enabled": True,
            "flatten_strategy": "cartesian",
            "defaults": {"item_purchase_type": "Catalogue", "minimum_quantity": 1},
            "price_export_mapping": {"direct_unit_cost": "installer"},
        },
    })


def _make_atlantic_file(tmp_path: Path, nb: int = 5, installer_price_offset: float = 0.0) -> Path:
    tmp_path.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atlantic 2026"
    ws["C4"] = "2026-01-01"
    ws["C5"] = "2026-12-31"
    ws.cell(row=9, column=2, value="Code article")
    ws.cell(row=9, column=3, value="Désignation")
    for i in range(nb):
        r = 10 + i
        ws.cell(row=r, column=2, value=f"CODE{i+1:03d}")
        ws.cell(row=r, column=3, value=f"Article {i+1}")
        ws.cell(row=r, column=4, value=1)
        ws.cell(row=r, column=5, value=float(100 + i * 10))
        ws.cell(row=r, column=6, value=float(90 + i * 10 + installer_price_offset))
    path = tmp_path / "atlantic.xlsx"
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# E2E 1 — Première ingestion Atlantic → tous CREATE → NEW_ARTICLE + NEW_ART_FRNS_CREATE
# ─────────────────────────────────────────────────────────────────────────────

def test_e2e_atlantic_premiere_ingestion(tmp_path: Path) -> None:
    path = _make_atlantic_file(tmp_path, nb=5)
    rule = _atlantic_rule()

    # Parsing
    result = parse_table_file(path, rule)
    assert len(result.products) == 5
    assert result.error_count == 0

    # Delta (aucun état connu → tout CREATE)
    delta = compute_delta(result.products, known_hashes={})
    assert len(delta.creates) == 5
    assert delta.total_changes == 5

    # Export Gery
    export_dir = tmp_path / "exports"
    gery = generate_gery_exports(delta, rule.gery_export, rule.supplier_code, export_dir,
                                  result.file_metadata.validity_start,
                                  result.file_metadata.validity_end)

    kinds = {f.kind for f in gery.files}
    assert "NEW_ARTICLE" in kinds
    assert "NEW_ART_FRNS_CREATE" in kinds
    assert "NEW_ART_FRNS_PRICE_UPDATE" not in kinds

    na = next(f for f in gery.files if f.kind == "NEW_ARTICLE")
    assert na.line_count == 5
    assert na.path.exists()


# ─────────────────────────────────────────────────────────────────────────────
# E2E 2 — Deuxième ingestion sans changement → zéro export
# ─────────────────────────────────────────────────────────────────────────────

def test_e2e_atlantic_idempotence(tmp_path: Path) -> None:
    path = _make_atlantic_file(tmp_path, nb=3)
    rule = _atlantic_rule()

    result1 = parse_table_file(path, rule)
    known_hashes = {p.supplier_product_code: compute_business_hash(p) for p in result1.products}

    # Même fichier, même résultat
    result2 = parse_table_file(path, rule)
    delta = compute_delta(result2.products, known_hashes=known_hashes)

    assert delta.unchanged == 3
    assert delta.total_changes == 0

    export_dir = tmp_path / "exports"
    gery = generate_gery_exports(delta, rule.gery_export, rule.supplier_code, export_dir)
    assert gery.files == []


# ─────────────────────────────────────────────────────────────────────────────
# E2E 3 — Mise à jour de prix → PRICE_CHANGE → NEW_ART_FRNS_PRICE_UPDATE
# ─────────────────────────────────────────────────────────────────────────────

def test_e2e_atlantic_price_update(tmp_path: Path) -> None:
    rule = _atlantic_rule()

    # Snapshot 1
    path1 = _make_atlantic_file(tmp_path / "v1", nb=3, installer_price_offset=0.0)
    result1 = parse_table_file(path1, rule)
    known_hashes = {p.supplier_product_code: compute_business_hash(p) for p in result1.products}

    # Snapshot 2 — prix installateur -5
    (tmp_path / "v2").mkdir()
    path2 = _make_atlantic_file(tmp_path / "v2", nb=3, installer_price_offset=-5.0)
    result2 = parse_table_file(path2, rule)
    delta = compute_delta(result2.products, known_hashes=known_hashes)

    assert len(delta.price_changes) == 3
    assert delta.total_changes == 3

    export_dir = tmp_path / "exports"
    gery = generate_gery_exports(delta, rule.gery_export, rule.supplier_code, export_dir)

    kinds = {f.kind for f in gery.files}
    assert "NEW_ART_FRNS_PRICE_UPDATE" in kinds
    assert "NEW_ARTICLE" not in kinds

    pu = next(f for f in gery.files if f.kind == "NEW_ART_FRNS_PRICE_UPDATE")
    assert pu.line_count == 3


# ─────────────────────────────────────────────────────────────────────────────
# E2E 4 — Produit supprimé → DELETE (pas d'export Gery, juste détection)
# ─────────────────────────────────────────────────────────────────────────────

def test_e2e_atlantic_delete_detection(tmp_path: Path) -> None:
    rule = _atlantic_rule()

    path1 = _make_atlantic_file(tmp_path / "v1", nb=5)
    result1 = parse_table_file(path1, rule)
    known_hashes = {p.supplier_product_code: compute_business_hash(p) for p in result1.products}

    (tmp_path / "v2").mkdir()
    path2 = _make_atlantic_file(tmp_path / "v2", nb=3)  # 2 produits disparus
    result2 = parse_table_file(path2, rule)
    delta = compute_delta(result2.products, known_hashes=known_hashes)

    assert len(delta.deletes) == 2
    delete_codes = {d.supplier_product_code for d in delta.deletes}
    assert "CODE004" in delete_codes
    assert "CODE005" in delete_codes


# ─────────────────────────────────────────────────────────────────────────────
# E2E 5 — Scénario mixte CREATE + PRICE_CHANGE + DELETE
# ─────────────────────────────────────────────────────────────────────────────

def test_e2e_atlantic_mixed_batch(tmp_path: Path) -> None:
    rule = _atlantic_rule()

    path1 = _make_atlantic_file(tmp_path / "v1", nb=4)
    result1 = parse_table_file(path1, rule)
    known_hashes = {p.supplier_product_code: compute_business_hash(p) for p in result1.products}

    # v2 : 3 inchangés + 1 disparu + 1 nouveau + prix modifiés
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atlantic 2026"
    ws["C4"] = "2026-01-01"
    ws["C5"] = "2026-12-31"
    ws.cell(row=9, column=2, value="Code")
    ws.cell(row=9, column=3, value="Désignation")
    # CODE001 → prix changé
    ws.cell(row=10, column=2, value="CODE001")
    ws.cell(row=10, column=3, value="Article 1")
    ws.cell(row=10, column=4, value=1)
    ws.cell(row=10, column=5, value=100.0)
    ws.cell(row=10, column=6, value=75.0)  # était 90
    # CODE002, CODE003 → inchangés
    for i in (1, 2):
        r = 11 + i - 1
        ws.cell(row=r, column=2, value=f"CODE{i+1:03d}")
        ws.cell(row=r, column=3, value=f"Article {i+1}")
        ws.cell(row=r, column=4, value=1)
        ws.cell(row=r, column=5, value=float(100 + i * 10))
        ws.cell(row=r, column=6, value=float(90 + i * 10))
    # CODE005 → nouveau
    ws.cell(row=13, column=2, value="CODE005")
    ws.cell(row=13, column=3, value="Article Nouveau")
    ws.cell(row=13, column=4, value=1)
    ws.cell(row=13, column=5, value=150.0)
    ws.cell(row=13, column=6, value=130.0)
    # CODE004 → absent → DELETE
    (tmp_path / "v2").mkdir()
    path2 = tmp_path / "v2" / "atlantic_v2.xlsx"
    wb.save(path2)

    result2 = parse_table_file(path2, rule)
    delta = compute_delta(result2.products, known_hashes=known_hashes)

    assert len(delta.creates) == 1    # CODE005
    assert len(delta.price_changes) == 1  # CODE001
    assert delta.unchanged == 2       # CODE002, CODE003
    assert len(delta.deletes) == 1    # CODE004


# ─────────────────────────────────────────────────────────────────────────────
# E2E 6 — business_hash stable (idempotence du hash)
# ─────────────────────────────────────────────────────────────────────────────

def test_e2e_business_hash_stable(tmp_path: Path) -> None:
    """Parser deux fois le même fichier Atlantic produit les mêmes hashes."""
    path = _make_atlantic_file(tmp_path, nb=5)
    rule = _atlantic_rule()

    r1 = parse_table_file(path, rule)
    r2 = parse_table_file(path, rule)

    hashes1 = sorted([compute_business_hash(p) for p in r1.products])
    hashes2 = sorted([compute_business_hash(p) for p in r2.products])
    assert hashes1 == hashes2


# ─────────────────────────────────────────────────────────────────────────────
# E2E 7 — Reactivation d'un produit supprimé
# ─────────────────────────────────────────────────────────────────────────────

def test_e2e_atlantic_reactivation(tmp_path: Path) -> None:
    rule = _atlantic_rule()

    path = _make_atlantic_file(tmp_path, nb=3)
    result = parse_table_file(path, rule)
    known_hashes = {p.supplier_product_code: compute_business_hash(p) for p in result.products}

    # CODE001 est marqué supprimé en base
    deleted = {"CODE001"}

    # Il réapparaît dans le nouveau fichier
    delta = compute_delta(result.products, known_hashes=known_hashes, deleted_codes=deleted)

    assert len(delta.reactivates) == 1
    assert delta.reactivates[0].supplier_product_code == "CODE001"
    assert delta.unchanged == 2
