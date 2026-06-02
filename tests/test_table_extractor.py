from __future__ import annotations

import io
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from middleware.parser.grammar import MappingRule
from middleware.parser.table_extractor import (
    _row_passes_filter,
    compute_business_hash,
    parse_table_file,
)
from middleware.parser.grammar import RowFilter
from middleware.parser.pivot import ProductPivot, PricePivot


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures — construction de fichiers Excel synthétiques en mémoire
# ─────────────────────────────────────────────────────────────────────────────

def _make_atlantic_xlsx(tmp_path: Path, nb_produits: int = 5) -> Path:
    """Crée un fichier Excel Atlantic synthétique avec nb_produits articles."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atlantic 2026"

    # Cartouche (lignes 1-8)
    ws["C2"] = "Toutes Sociétés Ramery"
    ws["C3"] = "Toute France"
    ws["C4"] = "2026-03-01"
    ws["C5"] = "2026-12-31"
    ws["E7"] = "Référence Atlantic 293157102 Validité 31/12/2026"

    # En-têtes ligne 9
    ws.cell(row=9, column=1, value="Base / Variante")
    ws.cell(row=9, column=2, value="Code article")
    ws.cell(row=9, column=3, value="Désignation")
    ws.cell(row=9, column=4, value="Quantité")
    ws.cell(row=9, column=5, value="Prix Public 2026")
    ws.cell(row=9, column=6, value="Prix installateur")

    # Données à partir de la ligne 10
    for i in range(nb_produits):
        row = 10 + i
        ws.cell(row=row, column=1, value="B")
        ws.cell(row=row, column=2, value=f"CODE{i+1:03d}")
        ws.cell(row=row, column=3, value=f"Article Test {i+1}")
        ws.cell(row=row, column=4, value=1)
        ws.cell(row=row, column=5, value=float(100 + i * 10))
        ws.cell(row=row, column=6, value=float(90 + i * 10))

    path = tmp_path / "atlantic_test.xlsx"
    wb.save(path)
    return path


def _make_atlantic_rule() -> MappingRule:
    """Construit la MappingRule Atlantic depuis un dict (sans YAML loader)."""
    return MappingRule.model_validate({
        "supplier_code": "atlantic_scga_chauffage",
        "mapping_version": 1,
        "sheet_match": "Atlantic 2026",
        "header_detection": {"mode": "explicit", "row": 9},
        "data_starts_row": 10,
        "extraction_mode": "table",
        "row_filter": {
            "must_have_value_in": ["B"],
            "exclude_if_starts_with": ["TARIF GROUPE"],
        },
        "columns": {
            "supplier_product_code": {"source_col": "B", "transform": ["strip", "to_uppercase"], "required": True},
            "designation": {"source_col": "C", "transform": "strip", "required": True},
            "family": {"constant": "Chauffage électrique"},
        },
        "prices": [
            {"type": "public", "source_col": "E", "transform": "parse_decimal_fr"},
            {"type": "installer", "source_col": "F", "transform": "parse_decimal_fr"},
        ],
        "attributes": [
            {"key": "quantity_pack", "source_col": "D", "data_type": "integer"},
            {"key": "base_variant", "source_col": "A", "data_type": "enum"},
        ],
        "file_metadata": {
            "validity_start": {"cell": "C4", "transform": "parse_date_iso"},
            "validity_end": {"cell": "C5", "transform": "parse_date_iso"},
            "contract_reference": {"regex": "Référence Atlantic (\\d+)", "in_cell": "E7"},
            "geographic_scope": {"cell": "C3"},
            "organizational_scope": {"cell": "C2"},
        },
        "gery_export": {
            "enabled": True,
            "flatten_strategy": "cartesian",
            "defaults": {"item_purchase_type": "Catalogue", "minimum_quantity": 1},
            "price_export_mapping": {"direct_unit_cost": "installer"},
        },
    })


# ─────────────────────────────────────────────────────────────────────────────
# Tests — filtre de lignes
# ─────────────────────────────────────────────────────────────────────────────

def test_row_filter_must_have_value() -> None:
    row_filter = RowFilter(must_have_value_in=["B"])
    # colonne B (index 1) vide → rejetée
    assert not _row_passes_filter(["B_val", None, "C_val"], row_filter)
    # colonne B non vide → acceptée
    assert _row_passes_filter(["A_val", "B_val", "C_val"], row_filter)


def test_row_filter_exclude_prefix() -> None:
    row_filter = RowFilter(exclude_if_starts_with=["TARIF GROUPE"])
    assert not _row_passes_filter(["TARIF GROUPE — Atlantic", "val", "val"], row_filter)
    assert _row_passes_filter(["341073", "val", "val"], row_filter)


def test_row_filter_must_have_any() -> None:
    row_filter = RowFilter(must_have_value_in_any=["G", "I"])
    # colonnes G (idx 6) et I (idx 8) toutes vides → rejetée
    row = [None] * 10
    assert not _row_passes_filter(row, row_filter)
    # colonne G non vide → acceptée
    row[6] = 5.01
    assert _row_passes_filter(row, row_filter)


# ─────────────────────────────────────────────────────────────────────────────
# Tests — parsing E2E avec fichier synthétique
# ─────────────────────────────────────────────────────────────────────────────

def test_parse_atlantic_nb_produits(tmp_path: Path) -> None:
    """Vérifie que le parseur extrait le bon nombre de produits."""
    path = _make_atlantic_xlsx(tmp_path, nb_produits=5)
    rule = _make_atlantic_rule()
    result = parse_table_file(path, rule)
    assert result.error_count == 0
    assert len(result.products) == 5


def test_parse_atlantic_champs_produit(tmp_path: Path) -> None:
    """Vérifie les champs du premier produit extrait."""
    path = _make_atlantic_xlsx(tmp_path, nb_produits=3)
    rule = _make_atlantic_rule()
    result = parse_table_file(path, rule)

    p = result.products[0]
    assert p.supplier_code == "atlantic_scga_chauffage"
    assert p.supplier_product_code == "CODE001"
    assert p.designation == "Article Test 1"
    assert p.family == "Chauffage électrique"
    assert p.product_kind == "physical"


def test_parse_atlantic_prix(tmp_path: Path) -> None:
    """Vérifie l'extraction des 2 prix par produit."""
    path = _make_atlantic_xlsx(tmp_path, nb_produits=3)
    rule = _make_atlantic_rule()
    result = parse_table_file(path, rule)

    p = result.products[0]
    assert len(p.prices) == 2
    types = {pr.price_type for pr in p.prices}
    assert types == {"public", "installer"}
    installer = next(pr for pr in p.prices if pr.price_type == "installer")
    assert installer.amount == Decimal("90")


def test_parse_atlantic_attributs(tmp_path: Path) -> None:
    """Vérifie l'extraction des attributs techniques."""
    path = _make_atlantic_xlsx(tmp_path, nb_produits=1)
    rule = _make_atlantic_rule()
    result = parse_table_file(path, rule)

    p = result.products[0]
    keys = {a.key for a in p.attributes}
    assert "quantity_pack" in keys
    assert "base_variant" in keys


def test_parse_atlantic_file_metadata(tmp_path: Path) -> None:
    """Vérifie l'extraction des métadonnées du cartouche."""
    path = _make_atlantic_xlsx(tmp_path, nb_produits=1)
    rule = _make_atlantic_rule()
    result = parse_table_file(path, rule)

    meta = result.file_metadata
    assert meta.validity_start is not None
    assert meta.validity_end is not None
    assert meta.contract_reference == "293157102"
    assert meta.geographic_scope == "Toute France"
    assert meta.organizational_scope == "Toutes Sociétés Ramery"


def test_parse_atlantic_ligne_vide_ignoree(tmp_path: Path) -> None:
    """Les lignes avec colonne B vide sont ignorées."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atlantic 2026"
    ws.cell(row=9, column=2, value="Code article")
    ws.cell(row=10, column=2, value="CODE001")
    ws.cell(row=10, column=3, value="Article valide")
    ws.cell(row=10, column=5, value=100.0)
    ws.cell(row=10, column=6, value=90.0)
    ws.cell(row=11, column=2, value=None)  # ligne vide → ignorée
    ws.cell(row=11, column=3, value="Article ignoré")
    path = tmp_path / "atlantic_vide.xlsx"
    wb.save(path)
    rule = _make_atlantic_rule()
    result = parse_table_file(path, rule)
    assert len(result.products) == 1


def test_parse_atlantic_idempotence(tmp_path: Path) -> None:
    """Parser deux fois le même fichier produit les mêmes business_hash."""
    path = _make_atlantic_xlsx(tmp_path, nb_produits=3)
    rule = _make_atlantic_rule()

    result1 = parse_table_file(path, rule)
    result2 = parse_table_file(path, rule)

    hashes1 = [compute_business_hash(p) for p in result1.products]
    hashes2 = [compute_business_hash(p) for p in result2.products]
    assert hashes1 == hashes2


def test_compute_business_hash_stable() -> None:
    """Le business_hash est stable quel que soit l'ordre des prix/attributs."""
    p = ProductPivot(
        supplier_code="test",
        supplier_product_code="CODE001",
        designation="Test Article",
        family="Chauffage",
        prices=[
            PricePivot(price_type="public", amount=Decimal("100")),
            PricePivot(price_type="installer", amount=Decimal("90")),
        ],
    )
    h1 = compute_business_hash(p)
    # Inverser l'ordre des prix — le hash doit rester identique
    p.prices.reverse()
    h2 = compute_business_hash(p)
    assert h1 == h2


# ─────────────────────────────────────────────────────────────────────────────
# Tests — transformations
# ─────────────────────────────────────────────────────────────────────────────

class TestTransforms:
    def test_parse_decimal_fr_virgule(self) -> None:
        from middleware.parser.transforms import _parse_decimal_fr
        assert _parse_decimal_fr("125,50") == Decimal("125.50")

    def test_parse_decimal_fr_espace(self) -> None:
        from middleware.parser.transforms import _parse_decimal_fr
        assert _parse_decimal_fr("1 250,00") == Decimal("1250.00")

    def test_parse_decimal_fr_invalide(self) -> None:
        from middleware.parser.transforms import _parse_decimal_fr
        with pytest.raises(ValueError):
            _parse_decimal_fr("A LA DEMANDE")

    def test_parse_date_iso_string(self) -> None:
        from middleware.parser.transforms import _parse_date_iso
        from datetime import date
        assert _parse_date_iso("2026-03-01") == date(2026, 3, 1)

    def test_parse_date_iso_datetime_object(self) -> None:
        from middleware.parser.transforms import _parse_date_iso
        from datetime import date, datetime
        assert _parse_date_iso(datetime(2026, 3, 1, 0, 0)) == date(2026, 3, 1)

    def test_parse_duration_fr(self) -> None:
        from middleware.parser.transforms import _parse_duration_fr
        assert _parse_duration_fr("4,33H") == Decimal("4.33")

    def test_col_letter_to_idx(self) -> None:
        from middleware.parser.transforms import col_letter_to_idx
        assert col_letter_to_idx("A") == 0
        assert col_letter_to_idx("B") == 1
        assert col_letter_to_idx("Z") == 25
        assert col_letter_to_idx("AA") == 26

    def test_cell_ref_to_row_col(self) -> None:
        from middleware.parser.transforms import cell_ref_to_row_col
        assert cell_ref_to_row_col("A1") == (0, 0)
        assert cell_ref_to_row_col("C4") == (3, 2)
        assert cell_ref_to_row_col("E7") == (6, 4)
