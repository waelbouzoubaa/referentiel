"""Tests de l'exporter Gery (NEW_ARTICLE).

Note : NEW_ART_FRNS_CREATE et NEW_ART_FRNS_PRICE_UPDATE (Phase 5) ne sont pas
encore implémentés — cf. tests xfail dans test_e2e_pipeline.py et
docs/DEPLOIEMENT.md (section "Points d'attention connus").
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from middleware.delta.engine import compute_delta
from middleware.exporter.gery import NEW_ARTICLE_COLS, _get_codes_with_prices, generate_gery_exports
from middleware.parser.grammar import GeryExportConfig
from middleware.parser.pivot import PricePivot, ProductPivot, VariantPivot

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _make_simple_product(code: str, installer: str = "90") -> ProductPivot:
    return ProductPivot(
        supplier_code="test",
        supplier_product_code=code,
        designation=f"Article {code}",
        family="Famille",
        prices=[
            PricePivot(price_type="public", amount=Decimal("100")),
            PricePivot(price_type="installer", amount=Decimal(installer)),
        ],
    )


def _make_matrix_product(code: str) -> ProductPivot:
    """Produit avec variantes (type Airisol)."""
    return ProductPivot(
        supplier_code="airisol",
        supplier_product_code=code,
        designation=f"Panneau {code}",
        family="Étanchéité",
        variants=[
            VariantPivot(
                variant_dimension="couleur",
                variant_value="ALU",
                variant_code="ALU",
                prices=[
                    PricePivot(price_type="list", amount=Decimal("10"),
                               tier_min_quantity=Decimal("0"), tier_max_quantity=Decimal("500")),
                    PricePivot(price_type="list", amount=Decimal("8"),
                               tier_min_quantity=Decimal("500"), tier_max_quantity=Decimal("1000")),
                ],
            ),
            VariantPivot(
                variant_dimension="couleur",
                variant_value="BLANC",
                variant_code="BLANC",
                prices=[
                    PricePivot(price_type="list", amount=Decimal("9"),
                               tier_min_quantity=Decimal("0"), tier_max_quantity=Decimal("500")),
                    PricePivot(price_type="list", amount=Decimal("7"),
                               tier_min_quantity=Decimal("500"), tier_max_quantity=Decimal("1000")),
                ],
            ),
        ],
    )


def _atlantic_config() -> GeryExportConfig:
    return GeryExportConfig(
        enabled=True,
        flatten_strategy="cartesian",
        defaults={"item_purchase_type": "Catalogue", "minimum_quantity": 1},
        price_export_mapping={"direct_unit_cost": "installer"},
    )


def _airisol_config() -> GeryExportConfig:
    return GeryExportConfig(
        enabled=True,
        flatten_strategy="cartesian",
        derived_code_template="{designation} | {variant_code} | {tier_label}",
        defaults={"item_purchase_type": "Catalogue", "minimum_quantity": 1},
        price_export_mapping={"direct_unit_cost": "list"},
    )


def _delta_with_creates(products: list[ProductPivot]):
    return compute_delta(products, known_hashes={})


# ─────────────────────────────────────────────────────────────────────────────
# Tests — gery_export désactivé
# ─────────────────────────────────────────────────────────────────────────────

def test_export_disabled_produces_no_files(tmp_path: Path) -> None:
    config = GeryExportConfig(enabled=False, blocked_reason="Test disabled")
    delta = _delta_with_creates([_make_simple_product("CODE001")])
    result = generate_gery_exports(delta, config, "test", tmp_path)
    assert result.files == []


# ─────────────────────────────────────────────────────────────────────────────
# Tests — CREATE → NEW_ARTICLE
# ─────────────────────────────────────────────────────────────────────────────

def test_creates_produce_new_article_file(tmp_path: Path) -> None:
    products = [_make_simple_product(f"CODE{i:03d}") for i in range(3)]
    delta = _delta_with_creates(products)
    result = generate_gery_exports(delta, _atlantic_config(), "atlantic", tmp_path)

    kinds = {f.kind for f in result.files}
    assert kinds == {"NEW_ARTICLE"}


def test_new_article_line_count(tmp_path: Path) -> None:
    products = [_make_simple_product(f"CODE{i:03d}") for i in range(5)]
    delta = _delta_with_creates(products)
    result = generate_gery_exports(delta, _atlantic_config(), "atlantic", tmp_path)

    na = next(f for f in result.files if f.kind == "NEW_ARTICLE")
    assert na.line_count == 5


def test_new_article_excel_content(tmp_path: Path) -> None:
    products = [_make_simple_product("CODE001")]
    delta = _delta_with_creates(products)
    result = generate_gery_exports(delta, _atlantic_config(), "atlantic", tmp_path)

    na = next(f for f in result.files if f.kind == "NEW_ARTICLE")
    wb = openpyxl.load_workbook(na.path)
    ws = wb.active

    header = [ws.cell(row=1, column=c).value for c in range(1, len(NEW_ARTICLE_COLS) + 1)]
    assert "Code article Frns" in header
    assert "Description" in header

    code_col = header.index("Code article Frns") + 1
    assert ws.cell(row=2, column=code_col).value == "CODE001"


def test_new_article_direct_unit_cost_uses_price_mapping(tmp_path: Path) -> None:
    products = [_make_simple_product("CODE001", installer="85")]
    delta = _delta_with_creates(products)
    result = generate_gery_exports(delta, _atlantic_config(), "atlantic", tmp_path)

    na = next(f for f in result.files if f.kind == "NEW_ARTICLE")
    wb = openpyxl.load_workbook(na.path)
    ws = wb.active

    header = [ws.cell(row=1, column=c).value for c in range(1, len(NEW_ARTICLE_COLS) + 1)]
    cost_col = header.index("Direct Unit Cost") + 1
    assert ws.cell(row=2, column=cost_col).value == pytest.approx(85.0)


# ─────────────────────────────────────────────────────────────────────────────
# Tests — codes article dérivés (produits simples et matriciels)
# ─────────────────────────────────────────────────────────────────────────────

def test_codes_with_prices_simple_product() -> None:
    p = _make_simple_product("CODE001")
    codes = _get_codes_with_prices(p, "installer", None)
    assert [code for code, _ in codes] == ["CODE001"]


def test_codes_with_prices_matrix_product() -> None:
    p = _make_matrix_product("PANN001")
    config = _airisol_config()
    codes = _get_codes_with_prices(
        p, config.price_export_mapping.direct_unit_cost, config.derived_code_template
    )

    # 2 variantes × 2 paliers = 4 codes
    rendered = [code for code, _ in codes]
    assert len(rendered) == 4
    assert any("ALU" in c for c in rendered)
    assert any("BLANC" in c for c in rendered)


def test_matrix_product_new_article_line_count(tmp_path: Path) -> None:
    """1 produit Airisol avec 2 variantes × 2 paliers = 4 lignes NEW_ARTICLE."""
    p = _make_matrix_product("PANN001")
    delta = compute_delta([p], known_hashes={})
    result = generate_gery_exports(delta, _airisol_config(), "airisol", tmp_path)

    na = next(f for f in result.files if f.kind == "NEW_ARTICLE")
    assert na.line_count == 4


# ─────────────────────────────────────────────────────────────────────────────
# Tests — fichier hash stable et fichier créé
# ─────────────────────────────────────────────────────────────────────────────

def test_generated_file_exists(tmp_path: Path) -> None:
    products = [_make_simple_product("CODE001")]
    delta = _delta_with_creates(products)
    result = generate_gery_exports(delta, _atlantic_config(), "atlantic", tmp_path)

    for f in result.files:
        assert f.path.exists()
        assert f.output_hash  # non-vide


def test_empty_delta_produces_no_files(tmp_path: Path) -> None:
    delta = compute_delta([], known_hashes={})
    result = generate_gery_exports(delta, _atlantic_config(), "atlantic", tmp_path)
    assert result.files == []


# ─────────────────────────────────────────────────────────────────────────────
# Tests — validité dates dans les exports
# ─────────────────────────────────────────────────────────────────────────────

def test_validity_dates_in_new_article(tmp_path: Path) -> None:
    products = [_make_simple_product("CODE001")]
    delta = _delta_with_creates(products)
    start = date(2026, 1, 1)
    end = date(2026, 12, 31)
    result = generate_gery_exports(delta, _atlantic_config(), "atlantic", tmp_path,
                                    validity_start=start, validity_end=end)

    na = next(f for f in result.files if f.kind == "NEW_ARTICLE")
    wb = openpyxl.load_workbook(na.path)
    ws = wb.active
    header = [ws.cell(row=1, column=c).value for c in range(1, len(NEW_ARTICLE_COLS) + 1)]

    deb_col = header.index("Starting Date") + 1
    assert ws.cell(row=2, column=deb_col).value == start.isoformat()
