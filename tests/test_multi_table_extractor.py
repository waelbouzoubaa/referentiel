from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from middleware.parser.grammar import MappingRule
from middleware.parser.multi_table_extractor import _render_template, _slugify, parse_multi_table_file


# ─────────────────────────────────────────────────────────────────────────────
# Fixture — fichier Agenor synthétique
# ─────────────────────────────────────────────────────────────────────────────

def _make_agenor_xlsx(tmp_path: Path) -> Path:
    """Crée un fichier Excel Agenor synthétique avec 2 sous-tableaux."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agenor 2026"

    # Cartouche
    ws["C2"] = "Validité de l'offre : 01/01/2026 au 31/12/2026"
    ws["A4"] = "Toute France"

    # ── Tableau 1 : entretien_bases_vie (lignes 8-18) ──────────────────────
    ws.cell(row=8, column=1, value="Taille base vie")
    ws.cell(row=8, column=2, value="Prix 1x/semaine")
    ws.cell(row=8, column=3, value="Durée max 1x")
    ws.cell(row=8, column=4, value="Prix 2x/semaine")
    ws.cell(row=8, column=5, value="Durée max 2x")
    ws.cell(row=8, column=6, value="Prix 5x/semaine")
    ws.cell(row=8, column=7, value="Durée max 5x")

    bases = [
        ("Petite (<20p)", 500.0, "2H", 800.0, "3,5H", 1500.0, "8H"),
        ("Moyenne (20-50p)", 700.0, "3H", 1100.0, "5H", 2000.0, "12H"),
        ("Grande (>50p)", 950.0, "4H", 1500.0, "7H", 2800.0, "16H"),
    ]
    for i, (taille, p1, d1, p2, d2, p5, d5) in enumerate(bases):
        r = 9 + i
        ws.cell(row=r, column=1, value=taille)
        ws.cell(row=r, column=2, value=p1)
        ws.cell(row=r, column=3, value=d1)
        ws.cell(row=r, column=4, value=p2)
        ws.cell(row=r, column=5, value=d2)
        ws.cell(row=r, column=6, value=p5)
        ws.cell(row=r, column=7, value=d5)

    # ── Tableau 2 : fournitures_consommables (lignes 23-27) ────────────────
    ws.cell(row=23, column=1, value="Tranche personnes")
    ws.cell(row=23, column=2, value="Forfait mensuel")

    tranches = [
        ("1-10 personnes", 120.0),
        ("11-25 personnes", 180.0),
        ("26-50 personnes", 260.0),
        (">50 personnes", 350.0),
    ]
    for i, (tranche, prix) in enumerate(tranches):
        r = 24 + i
        ws.cell(row=r, column=1, value=tranche)
        ws.cell(row=r, column=2, value=prix)

    path = tmp_path / "agenor_test.xlsx"
    wb.save(path)
    return path


def _make_agenor_rule() -> MappingRule:
    return MappingRule.model_validate({
        "supplier_code": "agenor",
        "mapping_version": 1,
        "sheet_match": "Agenor 2026",
        "header_detection": {"mode": "explicit", "row": 8},
        "data_starts_row": 9,
        "extraction_mode": "multi_table",
        "product_kind": "service",
        "tables": [
            {
                "name": "entretien_bases_vie",
                "description": "Forfait mensuel entretien",
                "zone": {"header_row": 8, "data_rows": "9:18", "cols": "A:G"},
                "layout": "matrix_2D",
                "col_dimensions": [
                    {"columns": ["B", "C"], "key": "frequency", "value": "1x_semaine", "price_col": "B", "max_time_col": "C"},
                    {"columns": ["D", "E"], "key": "frequency", "value": "2x_semaine", "price_col": "D", "max_time_col": "E"},
                    {"columns": ["F", "G"], "key": "frequency", "value": "5x_semaine", "price_col": "F", "max_time_col": "G"},
                ],
                "product_template": {
                    "designation_template": "Entretien base vie {taille_base_vie} — {frequency}",
                    "supplier_product_code_template": "AGEN-EBV-{taille_base_vie_slug}-{frequency}",
                    "family": "Entretien",
                    "subfamily": "Bases de vie",
                },
                "prices": [
                    {"type": "forfait", "source_col": "B", "transform": "parse_decimal_fr", "currency": "EUR"},
                ],
                "attributes": [
                    {"key": "max_monthly_time", "source_col": "C", "data_type": "duration", "unit": "h", "transform": "parse_duration_fr"},
                ],
            },
            {
                "name": "fournitures_consommables",
                "description": "Forfait mensuel fournitures",
                "zone": {"header_row": 23, "data_rows": "24:27", "cols": "A:B"},
                "layout": "barème_1D",
                "product_template": {
                    "designation_template": "Fournitures consommables sanitaires — {tranche_personnes}",
                    "supplier_product_code_template": "AGEN-FCS-{tranche_personnes_slug}",
                    "family": "Consommables",
                    "subfamily": "Sanitaires",
                },
                "prices": [
                    {"type": "forfait", "source_col": "B", "transform": "parse_decimal_fr", "currency": "EUR"},
                ],
                "attributes": [
                    {"key": "tranche_personnes", "source_col": "A", "data_type": "string"},
                ],
            },
        ],
        "file_metadata": {
            "validity_period": {
                "regex": "Validité de l'offre\\s*:\\s*(\\d{2}/\\d{2}/\\d{4})\\s*au\\s*(\\d{2}/\\d{2}/\\d{4})",
                "in_cell": "C2",
                "captures": {"validity_start": 1, "validity_end": 2},
                "transform": "parse_date_fr",
            },
            "geographic_scope": {"cell": "A4"},
        },
        "gery_export": {
            "enabled": False,
            "blocked_reason": "Modélisation des prestations à valider avec le métier",
        },
    })


# ─────────────────────────────────────────────────────────────────────────────
# Tests — utilitaires template
# ─────────────────────────────────────────────────────────────────────────────

def test_slugify_basic() -> None:
    assert _slugify("Petite (<20p)") == "PETITE_20P"


def test_slugify_spaces() -> None:
    assert _slugify("1-10 personnes") == "1_10_PERSONNES"


def test_render_template_row_only() -> None:
    context = {"__row_value__": "Tranche A", "__row_value_slug__": "TRANCHE_A"}
    result = _render_template("Code {tranche_personnes} — {tranche_personnes_slug}", context)
    assert result == "Code Tranche A — TRANCHE_A"


def test_render_template_row_and_col() -> None:
    context = {
        "__row_value__": "Petite",
        "__row_value_slug__": "PETITE",
        "frequency": "1x_semaine",
        "frequency_slug": "1X_SEMAINE",
    }
    result = _render_template("{taille_base_vie} — {frequency}", context)
    assert result == "Petite — 1x_semaine"


# ─────────────────────────────────────────────────────────────────────────────
# Tests — parsing E2E
# ─────────────────────────────────────────────────────────────────────────────

def test_parse_agenor_nb_produits_total(tmp_path: Path) -> None:
    """3 tailles × 3 fréquences + 4 tranches = 13 produits."""
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)
    assert result.error_count == 0
    assert len(result.products) == 13  # 3×3 + 4


def test_parse_agenor_product_kind(tmp_path: Path) -> None:
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)
    for p in result.products:
        assert p.product_kind == "service"


def test_parse_agenor_entretien_designation(tmp_path: Path) -> None:
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)

    # Produits du premier tableau
    entretien = [p for p in result.products if p.family == "Entretien"]
    assert len(entretien) == 9  # 3 tailles × 3 fréquences

    # Vérifie qu'une désignation est bien formée
    designations = {p.designation for p in entretien}
    assert any("1x_semaine" in d for d in designations)
    assert any("Petite" in d for d in designations)


def test_parse_agenor_entretien_code(tmp_path: Path) -> None:
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)

    entretien = [p for p in result.products if p.family == "Entretien"]
    codes = {p.supplier_product_code for p in entretien}
    assert any(c.startswith("AGEN-EBV-") for c in codes)


def test_parse_agenor_entretien_prix(tmp_path: Path) -> None:
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)

    entretien = [p for p in result.products if p.family == "Entretien"]
    for p in entretien:
        assert len(p.prices) == 1
        assert p.prices[0].price_type == "forfait"
        assert p.prices[0].amount > 0


def test_parse_agenor_consommables(tmp_path: Path) -> None:
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)

    consommables = [p for p in result.products if p.family == "Consommables"]
    assert len(consommables) == 4

    codes = {p.supplier_product_code for p in consommables}
    assert any(c.startswith("AGEN-FCS-") for c in codes)

    for p in consommables:
        assert len(p.prices) == 1
        assert p.prices[0].currency == "EUR"


def test_parse_agenor_consommables_prix_valeurs(tmp_path: Path) -> None:
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)

    consommables = sorted(
        [p for p in result.products if p.family == "Consommables"],
        key=lambda p: p.source_row or 0,
    )
    assert consommables[0].prices[0].amount == Decimal("120")
    assert consommables[3].prices[0].amount == Decimal("350")


def test_parse_agenor_file_metadata_captures(tmp_path: Path) -> None:
    """validity_period avec captures extrait validity_start et validity_end."""
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)

    meta = result.file_metadata
    assert meta.validity_start is not None
    assert meta.validity_end is not None
    from datetime import date
    assert meta.validity_start == date(2026, 1, 1)
    assert meta.validity_end == date(2026, 12, 31)


def test_parse_agenor_geographic_scope(tmp_path: Path) -> None:
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)
    assert result.file_metadata.geographic_scope == "Toute France"


def test_parse_agenor_attributs_entretien(tmp_path: Path) -> None:
    """Les attributs max_monthly_time sont extraits correctement."""
    path = _make_agenor_xlsx(tmp_path)
    rule = _make_agenor_rule()
    result = parse_multi_table_file(path, rule)

    entretien = [p for p in result.products if p.family == "Entretien"]
    for p in entretien:
        keys = {a.key for a in p.attributes}
        assert "max_monthly_time" in keys
