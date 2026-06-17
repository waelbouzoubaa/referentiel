from __future__ import annotations

from pathlib import Path
from typing import Any

from middleware.core.exceptions import ParsingError
from middleware.core.logging import get_logger

logger = get_logger(__name__)

Row = list[Any]
Sheet = list[Row]


def read_workbook(path: Path, sheet_name: str | None = None) -> dict[str, Sheet]:
    """Lit un fichier Excel et retourne un dict nom_feuille → lignes.

    Essaie calamine (rapide) d'abord, bascule sur openpyxl si indisponible.

    Args:
        path: Chemin vers le fichier .xlsx.
        sheet_name: Si fourni, ne charge que cette feuille.

    Returns:
        Dictionnaire {nom_feuille: [[valeurs], ...]}

    Raises:
        ParsingError: Si le fichier est illisible.
    """
    if not path.exists():
        raise ParsingError(f"Fichier introuvable : {path}", filename=path.name)

    try:
        return _read_with_calamine(path, sheet_name)
    except ImportError:
        logger.debug("calamine indisponible, bascule sur openpyxl")
        return _read_with_openpyxl(path, sheet_name)
    except Exception as exc:
        raise ParsingError(
            f"Erreur de lecture calamine ({path.name}) : {exc}",
            filename=path.name,
        ) from exc


def _read_with_calamine(path: Path, sheet_name: str | None) -> dict[str, Sheet]:
    """Lecture rapide via python-calamine."""
    from python_calamine import CalamineWorkbook

    wb = CalamineWorkbook.from_path(str(path))
    names = wb.sheet_names
    # Onglet demandé absent → on charge tout : find_sheet pourra matcher
    # (casse / auto) et surtout lister les vrais onglets en cas d'échec.
    if sheet_name is not None and sheet_name not in names:
        sheet_name = None
    sheets = {}
    for name in names:
        if sheet_name is not None and name != sheet_name:
            continue
        sheet = wb.get_sheet_by_name(name)
        sheets[name] = [list(row) for row in sheet.to_python(skip_empty_area=False)]
    return sheets


def _read_with_openpyxl(path: Path, sheet_name: str | None) -> dict[str, Sheet]:
    """Lecture via openpyxl (fallback, gère les cellules fusionnées)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    if sheet_name is not None and sheet_name not in names:
        sheet_name = None
    sheets = {}
    for name in names:
        if sheet_name is not None and name != sheet_name:
            continue
        ws = wb[name]
        rows: Sheet = []
        for row in ws.iter_rows(
            min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True
        ):
            rows.append(list(row))
        sheets[name] = rows
    wb.close()
    return sheets


def find_sheet(sheets: dict[str, Sheet], sheet_match: str | dict[str, str]) -> tuple[str, Sheet]:
    """Trouve la feuille correspondant à sheet_match.

    Args:
        sheets: Dict nom → lignes retourné par read_workbook.
        sheet_match: Nom exact, regex {'regex': '...'} ou 'auto'.

    Returns:
        Tuple (nom_feuille, lignes).

    Raises:
        ParsingError: Si aucune feuille ne correspond.
    """
    import re

    if sheet_match == "auto":
        # On prend la feuille avec le plus de lignes non vides
        best = max(
            sheets.items(),
            key=lambda kv: sum(1 for row in kv[1] if any(c is not None for c in row)),
        )
        return best

    if isinstance(sheet_match, str):
        if sheet_match in sheets:
            return sheet_match, sheets[sheet_match]
        # Tentative insensible à la casse
        for name, rows in sheets.items():
            if name.lower() == sheet_match.lower():
                return name, rows
        raise ParsingError(
            f"Feuille '{sheet_match}' introuvable. Feuilles disponibles : {list(sheets.keys())}"
        )

    if isinstance(sheet_match, dict) and "regex" in sheet_match:
        pattern = re.compile(sheet_match["regex"])
        for name, rows in sheets.items():
            if pattern.search(name):
                return name, rows
        raise ParsingError(
            f"Aucune feuille ne correspond au pattern '{sheet_match['regex']}'."
        )

    raise ParsingError(f"Format de sheet_match invalide : {sheet_match}")


def get_cell_value(sheet: Sheet, row_0: int, col_0: int) -> Any:
    """Retourne la valeur d'une cellule (0-indexed), None si hors limites."""
    if row_0 >= len(sheet):
        return None
    row = sheet[row_0]
    if col_0 >= len(row):
        return None
    return row[col_0]
