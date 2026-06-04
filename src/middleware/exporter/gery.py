from __future__ import annotations

import hashlib
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

from middleware.core.logging import get_logger
from middleware.delta.engine import ChangeType, DeltaResult, ProductDelta
from middleware.parser.grammar import GeryExportConfig
from middleware.parser.pivot import PricePivot, ProductPivot, VariantPivot

logger = get_logger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Résultat de génération
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class GeneratedFile:
    kind: str
    path: Path
    line_count: int
    output_hash: str


@dataclass
class GeryExportResult:
    files: list[GeneratedFile] = field(default_factory=list)
    generated_at: datetime = field(default_factory=datetime.utcnow)  # noqa: F841


# ─────────────────────────────────────────────────────────────────────────────
# Colonnes du fichier Gery (ordre imposé par l'ERP)
# ─────────────────────────────────────────────────────────────────────────────

NEW_ARTICLE_COLS = [
    "Code Fournisseur SAGE",       # null — alimenté depuis la base Gery à l'import
    "Code article Frns",
    "Description",
    "Article générique associé",
    "Gen. Prod. Posting Group",
    "Job Cost Code",
    "Tree Code",
    "Purchase Type",
    "Master Code",
    "Item Category Code",
    "Product Group Code",
    "Item Purchase Type",
    "Unité",
    "Code TVA",
    "Starting Date",
    "Minimum Quantity",
    "Direct Unit Cost",
    "Ending Date",
]


# ─────────────────────────────────────────────────────────────────────────────
# Point d'entrée principal
# ─────────────────────────────────────────────────────────────────────────────

def generate_gery_exports(
    delta: DeltaResult,
    export_config: GeryExportConfig,
    supplier_code: str,
    output_dir: Path,
    validity_start: date | None = None,
    validity_end: date | None = None,
) -> GeryExportResult:
    if not export_config.enabled:
        logger.info("gery_export désactivé", supplier_code=supplier_code)
        return GeryExportResult()

    output_dir.mkdir(parents=True, exist_ok=True)
    result = GeryExportResult()
    defaults = export_config.defaults
    price_field = export_config.price_export_mapping.direct_unit_cost

    # Créations + réactivations → NEW_ARTICLE
    rows = _build_rows(
        delta.creates + delta.reactivates,
        defaults,
        price_field,
        validity_start,
        validity_end,
    )

    if rows:
        path = output_dir / f"NEW_ARTICLE_{supplier_code}.xlsx"
        _write_excel(path, "NEW_ARTICLE", NEW_ARTICLE_COLS, rows)
        result.files.append(GeneratedFile(
            kind="NEW_ARTICLE",
            path=path,
            line_count=len(rows),
            output_hash=_file_hash(path),
        ))

    logger.info(
        "export Gery généré",
        supplier_code=supplier_code,
        lignes=len(rows),
    )
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Construction des lignes
# ─────────────────────────────────────────────────────────────────────────────

def _build_rows(
    deltas: list[ProductDelta],
    defaults: dict[str, Any],
    price_field: str,
    validity_start: date | None,
    validity_end: date | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for delta in deltas:
        if delta.new_product is None:
            continue
        p = delta.new_product
        for derived_code, price in _get_codes_with_prices(p, price_field):
            rows.append({
                "Code Fournisseur SAGE": None,
                "Code article Frns": derived_code,
                "Description": p.designation,
                "Article générique associé": defaults.get("article_generique", ""),
                "Gen. Prod. Posting Group": defaults.get("gen_prod_posting_group", ""),
                "Job Cost Code": defaults.get("job_cost_code", ""),
                "Tree Code": defaults.get("tree_code", ""),
                "Purchase Type": defaults.get("purchase_type", ""),
                "Master Code": defaults.get("master_code", ""),
                "Item Category Code": defaults.get("item_category_code", ""),
                "Product Group Code": defaults.get("product_group_code", ""),
                "Item Purchase Type": defaults.get("item_purchase_type", "Catalogue"),
                "Unité": defaults.get("unit_of_measure", "U"),
                "Code TVA": defaults.get("code_tva", "TVA20"),
                "Starting Date": validity_start,
                "Minimum Quantity": defaults.get("minimum_quantity", 1),
                "Direct Unit Cost": float(price.amount) if price else None,
                "Ending Date": validity_end,
            })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Flatten : produit simple ou matrice (variante × palier)
# ─────────────────────────────────────────────────────────────────────────────

def _get_codes_with_prices(
    product: ProductPivot,
    price_field: str,
) -> list[tuple[str, PricePivot | None]]:
    # Produit simple (table mode)
    if not product.variants:
        price = _find_price(product.prices, price_field)
        return [(product.supplier_product_code, price)]

    # Produit matrice (matrix mode) : une ligne par variante × palier
    result: list[tuple[str, PricePivot | None]] = []
    for t_idx, price in enumerate(
        p for v in product.variants for p in v.prices
        if not price_field or p.price_type == price_field
    ):
        variant = next(
            v for v in product.variants if price in v.prices
        )
        code = f"{product.supplier_product_code}-{variant.variant_code}-T{t_idx + 1}"
        result.append((code, price))

    if not result:
        for v_idx, variant in enumerate(product.variants):
            for t_idx, price in enumerate(variant.prices):
                code = f"{product.supplier_product_code}-{variant.variant_code}-T{t_idx + 1}"
                result.append((code, price))

    return result or [(product.supplier_product_code, None)]


def _find_price(prices: list[PricePivot], price_type: str) -> PricePivot | None:
    for p in prices:
        if p.price_type == price_type:
            return p
    return prices[0] if prices else None


# ─────────────────────────────────────────────────────────────────────────────
# Écriture Excel
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_excel(path: Path, sheet_name: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))

    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = max(len(col_name) + 2, 16)

    wb.save(path)


def _file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()
